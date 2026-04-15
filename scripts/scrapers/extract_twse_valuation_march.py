# -*- coding: utf-8 -*-
"""
從已解壓之證交所 C02001（3 月）Form1 工作表，擷取「當年 1–3 月」列之大盤
本益比 (P/E)、股價淨值比 (P/B)、殖利率 (%)。

輸出三個 .xlsx 檔（繁體欄位）。
"""
from __future__ import annotations

import re
import sys
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

PROJECT_DIR = Path(__file__).resolve().parents[2]
EXTRACTED = PROJECT_DIR / "data" / "twse_march_C02001" / "extracted"
OUT_DIR = PROJECT_DIR / "data" / "twse_march_C02001" / "output"
# 若 output 內 xlsx 被 Excel 鎖住無法覆寫，可改為 True 寫到 output_rebuild
OUT_FALLBACK = PROJECT_DIR / "data" / "twse_march_C02001" / "output_rebuild"


def find_form1_xls(folder: Path) -> Path | None:
    for pat in ("*.xls", "*.XLS"):
        xs = list(folder.glob(pat))
        if xs:
            return xs[0]
    return None


def extract_row(form1: pd.DataFrame) -> tuple[str, float, float, float] | None:
    """回傳 (期別標籤, PE, 殖利率%, P/B)。"""
    c0 = form1[0].astype(str)
    # 當年 1–3 月彙總列（證交所月報慣用標示）
    mask = c0.str.contains(r"1[-–]3", na=False, regex=True)
    if not mask.any():
        mask = c0.str.contains("1-3", na=False)
    if not mask.any():
        return None
    idx = mask[mask].index[0]
    pe = form1.iloc[idx, 11]
    dy = form1.iloc[idx, 12]
    pb = form1.iloc[idx, 13]
    label = str(form1.iloc[idx, 0]).strip()
    return label, float(pe), float(dy), float(pb)


def main() -> None:
    if not EXTRACTED.is_dir():
        print("找不到解壓目錄：", EXTRACTED, file=sys.stderr)
        sys.exit(1)

    rows: list[dict] = []
    for folder in sorted(EXTRACTED.iterdir()):
        if not folder.is_dir():
            continue
        m = re.match(r"^(\d{4})_03$", folder.name)
        if not m:
            continue
        year = int(m.group(1))
        xls = find_form1_xls(folder)
        if not xls:
            rows.append(
                {
                    "西元年": year,
                    "期別標籤": None,
                    "數值": None,
                    "狀態": "找不到 xls",
                    "檔案": None,
                }
            )
            continue
        try:
            form1 = pd.read_excel(xls, sheet_name="Form1", header=None)
        except Exception as e:
            rows.append(
                {
                    "西元年": year,
                    "期別標籤": None,
                    "數值": None,
                    "狀態": f"讀取失敗: {e}",
                    "檔案": xls.name,
                }
            )
            continue
        got = extract_row(form1)
        if not got:
            rows.append(
                {
                    "西元年": year,
                    "期別標籤": None,
                    "數值": None,
                    "狀態": "找不到 1-3 月列",
                    "檔案": xls.name,
                }
            )
            continue
        label, pe, dy, pb = got
        rows.append(
            {
                "西元年": year,
                "期別標籤": label,
                "本益比_PE_倍": pe,
                "殖利率_百分比": dy,
                "股價淨值比_PB_倍": pb,
                "狀態": "ok",
                "檔案": xls.name,
            }
        )

    df = pd.DataFrame(rows)
    OUT_DIR.mkdir(parents=True, exist_ok=True)

    note = (
        "資料來源：臺灣證券交易所「市場交易月報」報表 C02001（證券市場統計概要與市場總市值、"
        "投資報酬率、本益比、殖利率一覽表），Form1 工作表「當年 1–3 月」彙總列；"
        "欄位對應表頭 P/E Ratio (PER)、Dividend Yield (%)、P/B Ratio (PBR)。"
    )

    def write_one(path: Path, value_col: str, unit: str) -> None:
        out = df[["西元年", "期別標籤", value_col, "狀態", "檔案"]].copy()
        out = out.rename(columns={value_col: f"數值_{unit}"})
        # 先寫「資料」：Excel 預設顯示第一個工作表，避免使用者只看到「說明」以為沒資料
        with pd.ExcelWriter(path, engine="openpyxl") as w:
            out.to_excel(w, sheet_name="資料", index=False)
            pd.DataFrame([{"說明": note}]).to_excel(
                w, sheet_name="說明", index=False
            )
        wb = load_workbook(path)
        wb.active = wb.sheetnames.index("資料")
        wb["資料"].freeze_panes = "A2"
        wb.save(path)

    def try_write_all(target: Path) -> bool:
        try:
            write_one(
                target / "本益比_PE_大盤_三月彙總_2006-2026.xlsx",
                "本益比_PE_倍",
                "倍",
            )
            write_one(
                target / "股價淨值比_PB_大盤_三月彙總_2006-2026.xlsx",
                "股價淨值比_PB_倍",
                "倍",
            )
            write_one(
                target / "殖利率_大盤_三月彙總_2006-2026.xlsx",
                "殖利率_百分比",
                "百分比",
            )
            return True
        except PermissionError:
            return False

    out_dir = OUT_DIR
    if not try_write_all(out_dir):
        OUT_FALLBACK.mkdir(parents=True, exist_ok=True)
        if try_write_all(OUT_FALLBACK):
            out_dir = OUT_FALLBACK
            print(
                "注意：原 output 資料夾內檔案可能被 Excel 鎖住，已改寫入：",
                OUT_FALLBACK,
                file=sys.stderr,
            )
        else:
            print(
                "無法寫入 Excel（請關閉已開啟的 .xlsx 後再執行）。",
                file=sys.stderr,
            )
            sys.exit(1)

    ok = (df["狀態"] == "ok").sum()
    print(f"已寫入 {out_dir}")
    print(f"成功筆數：{ok} / {len(df)}")
    print("檔案：")
    for name in (
        "本益比_PE_大盤_三月彙總_2006-2026.xlsx",
        "股價淨值比_PB_大盤_三月彙總_2006-2026.xlsx",
        "殖利率_大盤_三月彙總_2006-2026.xlsx",
    ):
        print(" ", out_dir / name)


if __name__ == "__main__":
    main()
