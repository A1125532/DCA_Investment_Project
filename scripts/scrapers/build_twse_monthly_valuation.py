# -*- coding: utf-8 -*-
"""
下載證交所 C02001 月報（每年每月），從 Form1 擷取大盤本益比、殖利率、股價淨值比。

優先比對列標「民國年(1-M)月」（與該月統計一致）；若無則嘗試「民國年  M月」、最後「M月」。

輸出三份 xlsx（資料工作表置於第一頁）。
"""
from __future__ import annotations

import argparse
import io
import json
import re
import sys
import time
import urllib.error
import urllib.request
import zipfile
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

PROJECT_DIR = Path(__file__).resolve().parents[2]
ZIP_DIR = PROJECT_DIR / "data" / "twse_c02001_monthly" / "zips"
META_PATH = PROJECT_DIR / "data" / "twse_c02001_monthly" / "download_log.jsonl"
OUT_DIR = PROJECT_DIR / "data" / "twse_c02001_monthly" / "output"
BASE_URL = "https://www.twse.com.tw/staticFiles/inspection/inspection/02/001"
USER_AGENT = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
)


def read_form1_from_zip(zpath: Path) -> pd.DataFrame:
    with zipfile.ZipFile(zpath, "r") as zf:
        names = [n for n in zf.namelist() if n.lower().endswith(".xls")]
        if not names:
            raise ValueError("zip 內無 xls")
        raw = zf.read(names[0])
    return pd.read_excel(io.BytesIO(raw), sheet_name="Form1", header=None)


def row_has_pe(df: pd.DataFrame, i: int) -> bool:
    v = df.iloc[i, 11]
    return pd.notna(v) and isinstance(v, (int, float))


def extract_valuation(
    form1: pd.DataFrame, year: int, month: int
) -> tuple[str, float, float, float] | None:
    roc = year - 1911
    col0 = form1[0].astype(str)

    pat_ytd = re.compile(rf"{roc}\s*年\s*\(\s*1\s*-\s*{month}\s*\)\s*月")
    for i in range(len(col0)):
        if pat_ytd.search(col0.iloc[i]) and row_has_pe(form1, i):
            s = col0.iloc[i].strip()
            return (
                s,
                float(form1.iloc[i, 11]),
                float(form1.iloc[i, 12]),
                float(form1.iloc[i, 13]),
            )

    pat_roc_m = re.compile(rf"^{roc}\s*年\s*{month}\s*月\s*$")
    for i in range(len(col0)):
        if pat_roc_m.match(col0.iloc[i].strip()) and row_has_pe(form1, i):
            s = col0.iloc[i].strip()
            return (
                s,
                float(form1.iloc[i, 11]),
                float(form1.iloc[i, 12]),
                float(form1.iloc[i, 13]),
            )

    target = f"{month}月"
    last_i: int | None = None
    for i in range(len(col0)):
        if col0.iloc[i].strip() == target and row_has_pe(form1, i):
            last_i = i
    if last_i is not None:
        s = col0.iloc[last_i].strip()
        return (
            s,
            float(form1.iloc[last_i, 11]),
            float(form1.iloc[last_i, 12]),
            float(form1.iloc[last_i, 13]),
        )
    return None


def download_zip(yyyymm: str, dest: Path, timeout: int = 120) -> str:
    url = f"{BASE_URL}/{yyyymm}_C02001.zip"
    last_err = ""
    for attempt in range(3):
        req = urllib.request.Request(url, headers={"User-Agent": USER_AGENT})
        try:
            with urllib.request.urlopen(req, timeout=timeout) as r:
                data = r.read()
            break
        except urllib.error.HTTPError as e:
            if e.code == 404:
                return "404"
            last_err = f"http_{e.code}"
        except urllib.error.URLError as e:
            last_err = f"URLError:{e.reason!s}"
        except Exception as e:
            last_err = type(e).__name__
        time.sleep(2.0 * (attempt + 1))
    else:
        return last_err or "download_fail"
    dest.parent.mkdir(parents=True, exist_ok=True)
    dest.write_bytes(data)
    if not zipfile.is_zipfile(dest):
        try:
            dest.unlink()
        except OSError:
            pass
        return "not_zip"
    return "ok"


def write_three_excels(df: pd.DataFrame, out_dir: Path) -> None:
    out_dir.mkdir(parents=True, exist_ok=True)
    note = (
        "資料來源：臺灣證券交易所「市場交易月報」C02001，Form1；"
        "擷取規則：優先「民國年(1-當月)月」列（與該月大盤 P/E、殖利率、P/B 一致），"
        "備援「民國年 當月」列或單獨「M月」列。欄位對應表頭 PER、Dividend Yield(%)、PBR。"
    )

    def write_one(path: Path, value_col: str, unit: str) -> None:
        sub = df[
            ["西元年", "月份", "期別標籤", value_col, "狀態", "檔案"]
        ].copy()
        sub = sub.rename(columns={value_col: f"數值_{unit}"})
        with pd.ExcelWriter(path, engine="openpyxl") as w:
            sub.to_excel(w, sheet_name="資料", index=False)
            pd.DataFrame([{"說明": note}]).to_excel(
                w, sheet_name="說明", index=False
            )
        wb = load_workbook(path)
        wb.active = wb.sheetnames.index("資料")
        wb["資料"].freeze_panes = "A2"
        wb.save(path)

    write_one(
        out_dir / "本益比_PE_大盤_每月_2006-2026.xlsx",
        "本益比_PE_倍",
        "倍",
    )
    write_one(
        out_dir / "股價淨值比_PB_大盤_每月_2006-2026.xlsx",
        "股價淨值比_PB_倍",
        "倍",
    )
    write_one(
        out_dir / "殖利率_大盤_每月_2006-2026.xlsx",
        "殖利率_百分比",
        "百分比",
    )


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--start-year", type=int, default=2006)
    ap.add_argument("--end-year", type=int, default=2026)
    ap.add_argument("--sleep", type=float, default=3.0, help="每次下載間隔秒數")
    ap.add_argument("--skip-download", action="store_true", help="只解析已存在 zip")
    args = ap.parse_args()

    rows: list[dict] = []
    ZIP_DIR.mkdir(parents=True, exist_ok=True)
    if META_PATH.exists():
        META_PATH.unlink()

    for y in range(args.start_year, args.end_year + 1):
        for m in range(1, 13):
            yyyymm = f"{y}{m:02d}"
            zpath = ZIP_DIR / f"{yyyymm}_C02001.zip"
            log: dict = {"yyyymm": yyyymm, "year": y, "month": m}

            if zpath.is_file() and not zipfile.is_zipfile(zpath):
                try:
                    zpath.unlink()
                except OSError:
                    pass

            if not args.skip_download:
                if not zpath.is_file():
                    time.sleep(args.sleep)
                    st = download_zip(yyyymm, zpath)
                    log["download"] = st
                    if st != "ok":
                        rows.append(
                            {
                                "西元年": y,
                                "月份": m,
                                "期別標籤": None,
                                "本益比_PE_倍": None,
                                "殖利率_百分比": None,
                                "股價淨值比_PB_倍": None,
                                "狀態": f"下載失敗:{st}",
                                "檔案": None,
                            }
                        )
                        with META_PATH.open("a", encoding="utf-8") as f:
                            f.write(json.dumps(log, ensure_ascii=False) + "\n")
                        continue
                else:
                    log["download"] = "cached"
            elif not zpath.is_file():
                rows.append(
                    {
                        "西元年": y,
                        "月份": m,
                        "期別標籤": None,
                        "本益比_PE_倍": None,
                        "殖利率_百分比": None,
                        "股價淨值比_PB_倍": None,
                        "狀態": "無 zip（略過下載）",
                        "檔案": None,
                    }
                )
                continue

            try:
                form1 = read_form1_from_zip(zpath)
            except Exception as e:
                rows.append(
                    {
                        "西元年": y,
                        "月份": m,
                        "期別標籤": None,
                        "本益比_PE_倍": None,
                        "殖利率_百分比": None,
                        "股價淨值比_PB_倍": None,
                        "狀態": f"讀檔失敗:{e}",
                        "檔案": zpath.name,
                    }
                )
                log["parse"] = str(e)
                with META_PATH.open("a", encoding="utf-8") as f:
                    f.write(json.dumps(log, ensure_ascii=False) + "\n")
                continue

            got = extract_valuation(form1, y, m)
            if not got:
                rows.append(
                    {
                        "西元年": y,
                        "月份": m,
                        "期別標籤": None,
                        "本益比_PE_倍": None,
                        "殖利率_百分比": None,
                        "股價淨值比_PB_倍": None,
                        "狀態": "找不到對應月份列",
                        "檔案": zpath.name,
                    }
                )
                log["parse"] = "no_row"
            else:
                label, pe, dy, pb = got
                rows.append(
                    {
                        "西元年": y,
                        "月份": m,
                        "期別標籤": label,
                        "本益比_PE_倍": pe,
                        "殖利率_百分比": dy,
                        "股價淨值比_PB_倍": pb,
                        "狀態": "ok",
                        "檔案": zpath.name,
                    }
                )
                log["parse"] = "ok"
            with META_PATH.open("a", encoding="utf-8") as f:
                f.write(json.dumps(log, ensure_ascii=False) + "\n")

    out_df = pd.DataFrame(rows)
    try:
        write_three_excels(out_df, OUT_DIR)
        out_msg = str(OUT_DIR)
    except PermissionError:
        alt = PROJECT_DIR / "data" / "twse_c02001_monthly" / "output_rebuild"
        alt.mkdir(parents=True, exist_ok=True)
        write_three_excels(out_df, alt)
        out_msg = str(alt) + "（原 output 可能被 Excel 鎖住）"

    ok = (out_df["狀態"] == "ok").sum()
    print(f"完成。成功 {ok} / {len(out_df)} 筆。輸出：{out_msg}")


if __name__ == "__main__":
    main()
